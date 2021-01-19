import openpyxl
from pathlib import Path
import getpass

desktop_path = '/home/' + getpass.getuser() + '/Desktop/'

class ExcelParser(object):
    def __init__(self, file_name, output_name, threshold):
        self.file_name = file_name
        self.output_name = output_name
        # if threshold <= 0, assume no threshold (i.e. treat the whole file as one page)
        self.threshold = threshold

    def load_file(self):
        """
        Saves input Excel file as wb_obj and sheet to be read later on.
        """
        # Setting the path to the xlsx file:
        # xlsx_file = Path(self.file_name)

        # read the excel file
        wb_obj = openpyxl.load_workbook(desktop_path + self.file_name)

        # read the active sheet from the excel file
        sheet = wb_obj.active

        return wb_obj, sheet

    def build_data(self, sheet):
        """
        Returns contents of active sheet in the form of a dictionary.
        
        Args:
            sheet: the active sheet corresponding to input Excel file

        Returns:
            data (dict): key (str) = column title, value = contents of that column
        """
        # note: row 1 is just titles
        col_names = []
        for col in sheet.iter_cols(1, sheet.max_column):
            col_names.append(col[0].value)
        # the last column is the Total (quantity * wholesale price), so rename
        col_names[-1] = 'Total'

        # read file contents into a python dictionary
        data = {}

        for name in col_names:
            data[name] = []

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue
            else:
                data[col_names[0]].append(row[0])
                data[col_names[1]].append(row[1])
                data[col_names[2]].append(row[2])
                data[col_names[3]].append(row[3][2:-1])
                data[col_names[4]].append(row[4])
                data[col_names[5]].append(row[5])
                data[col_names[6]].append(row[6])
        
        return data



    def calculate_totals(self, item_num, quantity, price):
        """
        Returns calculated totals by every threshold-number of rows.
        
        Args: (value at index i of each list should correspond to the same item)
            item_num (list): list of item numbers
            quantity (list): list of quantities
            price (list): list of prices
        
        Returns:
            totals (dict): key (string) = item number, value (list) = [quantity, price, quantity*price]
                corresponding to that item
        """
        # make sure item_num, quantity, and price should have the same length
        if not ( len(item_num) == len(quantity) and len(quantity) == len(price) ):
            raise Exception("Error: the list inputs do not all have the same length.")

        totals = {}
        total_rows = len(item_num)

        for i in range(total_rows):
            # if this is the first time item_num[i] is seen
            if totals.get(item_num[i]) == None:
                totals[item_num[i]] = [quantity[i], price[i], quantity[i] * price[i]]
            else:
                # temp variables for current quantity, price, total
                temp_quantity, temp_price, temp_total = quantity[i], price[i], quantity[i] * price[i]

                # increment quantity
                totals[item_num[i]][0] += temp_quantity
                # price stays the same
                
                totals[item_num[i]][1] = temp_price
                # update total
                totals[item_num[i]][2] += temp_quantity * temp_price
        
        return totals

    def final_receipt(self, data):
        """
        Returns a list of calculated totals.

        Args:
            data (dict): contents of the input Excel file in the form of a dictionary
        
        Returns:
            receipts (list of dictionaries): list of totals (dict) for each page.
                The length of receipts is number of pages,
                where each pages has at most the threshold-number of rows.
        """
        threshold = self.threshold - 1  # because 0-index
        total_rows = len(data.get('PO Number'))
        finished = False
        start = 0
        end = threshold - 1  # because first row is just titles
        receipts = []

        while True:
            if end < total_rows:
                curr_totals = self.calculate_totals(data.get('Item Number')[start:end], data.get('Quantity')[start:end], data.get('Wholesale Price')[start:end])
            # on the last page
            else:
                finished = True
                curr_totals = self.calculate_totals(data.get('Item Number')[start:], data.get('Quantity')[start:], data.get('Wholesale Price')[start:])
            
            receipts += [curr_totals]

            # increment start and end
            start = end
            end += threshold

            if finished:
                break
        
        return receipts

    def write_cell(self, sheet, row_dest, col_dest, val):
        """
        Changes cell value in Excel sheet.
        Note: row is 1-indexed

        Args:
            sheet: Excel sheet object
            row_dest: row of cell to be changed
            col_dest: column of cell to be changed
            val: new value of that cell
        """
        c = sheet.cell(row = row_dest, column = col_dest)
        c.value = val

    def output_file(self, wb_obj, sheet, data):
        """
        Duplicates the original file to now include calculated receipts for every page
        and saves this as a new file.

        Args:
            wb_obj: workbook object corresponding to the input Excel file
            sheet: the active sheet corresponding to wb_obj
            data (dict): contents of the input Excel file in the form of a dictionary
        """
        threshold = self.threshold
        total_rows = len(data.get('PO Number'))  # not including title row
        c1 = 8  # item number
        c2 = 9  # <quantity> X <price> = <quantity*price>

        # assume no threshold, so don't call final_receipt()
        if threshold <= 0:
            # totals is a dictionary object
            totals = self.calculate_totals(data.get('Item Number'), data.get('Quantity'), data.get('Wholesale Price'))
            # + 1 because write_cell() uses 1-index
            # + 1 again because we want the last row to have something written
            curr_row = total_rows - len(totals) + 1 + 1

            for key in totals:
                q, p, t = totals.get(key)
                self.write_cell(sheet, curr_row, c1, key)
                self.write_cell(sheet, curr_row, c2, str(q) + ' X ' + str(p) + ' = ' + str(t))
                curr_row += 1
        else:
            curr_row = 0  # initialize
            r = self.final_receipt(data)

            for i, curr_receipt in enumerate(r):
                if i == 0:  # if first time, initialize curr_row
                    curr_row = threshold - len(curr_receipt)
                elif i == len(r) - 1:  # if at last receipt, we know curr_row
                    # + 1 because write_cell() uses 1-index
                    # + 1 again because we want the last row to have something written
                    curr_row = total_rows - len(curr_receipt) + 1 + 1
                else:
                    curr_row += threshold - len(curr_receipt) - 1

                for key in curr_receipt:
                    q, p, t = curr_receipt.get(key)
                    self.write_cell(sheet, curr_row, c1, key)
                    self.write_cell(sheet, curr_row, c2, str(q) + ' X ' + str(p) + ' = ' + str(t))
                    curr_row += 1

        # save excel file
        wb_obj.save(desktop_path + self.output_name)

    def main(self):
        wb_obj, sheet = self.load_file()
        data = self.build_data(sheet)
        self.output_file(wb_obj, sheet, data)
        print(self.file_name)
